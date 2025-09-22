from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.core.paginator import Paginator
from django.db.models import Q
from django.utils import timezone
from datetime import datetime, timedelta
import json
from .models import Patient
from appointments.models import Appointment
# Temporarily disabled medical records imports until migrations are fixed
# from medical_records.models import MedicalRecord, Prescription, LabResult
from doctors.models import Doctor


@login_required
def patient_dashboard(request):
    """Patient dashboard view"""
    if request.user.user_type != 'patient':
        messages.error(request, 'Access denied.')
        return redirect('healthcare_project:home')
    
    patient = get_object_or_404(Patient, user=request.user)
    today = timezone.now().date()
    
    # Get dashboard statistics
    total_appointments = Appointment.objects.filter(patient=patient).count()
    upcoming_appointments_count = Appointment.objects.filter(
        patient=patient,
        appointment_date__gte=today,
        status__in=['scheduled', 'confirmed', 'pending']
    ).count()
    
    # Get pending and cancelled appointment counts
    pending_appointments_count = Appointment.objects.filter(
        patient=patient,
        status='pending'
    ).count()
    
    cancelled_appointments_count = Appointment.objects.filter(
        patient=patient,
        status='cancelled'
    ).count()
    
    # Get upcoming appointments for display
    upcoming_appointments = Appointment.objects.filter(
        patient=patient,
        appointment_date__gte=today,
        status__in=['scheduled', 'confirmed', 'pending']
    ).order_by('appointment_date', 'appointment_time')[:5]
    
    # Get next appointment
    next_appointment = Appointment.objects.filter(
        patient=patient,
        appointment_date__gte=today,
        status__in=['scheduled', 'confirmed', 'pending']
    ).order_by('appointment_date', 'appointment_time').first()
    
    # Get recent medical records (temporarily disabled)
    recent_records = []  # MedicalRecord.objects.filter(patient=patient).order_by('-date_created')[:5]
    
    # Mock data for enhanced dashboard display
    active_prescriptions = []  # Will be replaced when medical_records is fixed
    recent_lab_results = []  # Will be replaced when medical_records is fixed
    
    # Get active prescriptions count (temporarily using mock data)
    active_prescriptions_count = 3  # Mock data - will be replaced when medical_records is fixed
    
    # Get recent lab results count (temporarily using mock data)
    recent_lab_results_count = 2  # Mock data - will be replaced when medical_records is fixed
    
    # Get recent completed appointments
    completed_appointments = Appointment.objects.filter(
        patient=patient,
        status='completed'
    ).order_by('-appointment_date', '-appointment_time')[:5]

    context = {
        'patient': patient,
        'total_appointments': total_appointments,
        'upcoming_appointments': upcoming_appointments,
        'upcoming_appointments_count': upcoming_appointments_count,
        'pending_appointments_count': pending_appointments_count,
        'cancelled_appointments_count': cancelled_appointments_count,
        'completed_appointments': completed_appointments,
        'next_appointment': next_appointment,
        'recent_records': recent_records,
        'active_prescriptions': active_prescriptions,
        'active_prescriptions_count': active_prescriptions_count,
        'recent_lab_results': recent_lab_results,
        'recent_lab_results_count': recent_lab_results_count,
    }
    return render(request, 'patients/dashboard_enhanced.html', context)


@login_required
def patient_profile(request):
    """Patient profile view"""
    if request.user.user_type != 'patient':
        messages.error(request, 'Access denied.')
        return redirect('healthcare_project:home')
    
    patient = get_object_or_404(Patient, user=request.user)
    
    if request.method == 'POST':
        # Handle profile update
        patient.date_of_birth = request.POST.get('date_of_birth', patient.date_of_birth)
        patient.gender = request.POST.get('gender', patient.gender)
        patient.blood_type = request.POST.get('blood_type', patient.blood_type)
        patient.address = request.POST.get('address', patient.address)
        patient.emergency_contact_name = request.POST.get('emergency_contact_name', patient.emergency_contact_name)
        patient.emergency_contact_phone = request.POST.get('emergency_contact_phone', patient.emergency_contact_phone)
        patient.save()
        
        request.user.first_name = request.POST.get('first_name', request.user.first_name)
        request.user.last_name = request.POST.get('last_name', request.user.last_name)
        request.user.email = request.POST.get('email', request.user.email)
        request.user.phone_number = request.POST.get('phone_number', request.user.phone_number)
        request.user.save()
        
        messages.success(request, 'Profile updated successfully.')
        return redirect('patients:profile')
    
    context = {
        'patient': patient,
    }
    return render(request, 'patients/profile.html', context)


@login_required
def patient_appointments(request):
    """Patient appointments view"""
    if request.user.user_type != 'patient':
        messages.error(request, 'Access denied.')
        return redirect('healthcare_project:home')
    
    patient = get_object_or_404(Patient, user=request.user)
    today = timezone.now().date()
    now = timezone.now()
    
    # Get all appointments for this patient only
    appointments = Appointment.objects.filter(patient=patient).order_by('-appointment_date', '-appointment_time')
    
    # Separate by status - non-overlapping filters
    upcoming_appointments = appointments.filter(
        Q(appointment_date__gt=today) |
        Q(appointment_date=today, appointment_time__gte=now.time()),
        status__in=['scheduled', 'confirmed']
    ).order_by('appointment_date', 'appointment_time')
    
    # Pending should only be past scheduled appointments or specifically marked as pending
    pending_appointments = appointments.filter(
        Q(appointment_date__lt=today, status='scheduled') |
        Q(status='pending')
    )
    
    # Filter completed and cancelled appointments
    completed_appointments = appointments.filter(status='completed').order_by('-appointment_date', '-appointment_time')
    cancelled_appointments = appointments.filter(status='cancelled')
    
    # Get next appointment for alert
    next_appointment = upcoming_appointments.first()
    
    # Pagination for completed appointments
    paginator = Paginator(completed_appointments, 10)
    page_number = request.GET.get('page')
    completed_page_obj = paginator.get_page(page_number)
    
    # Get appointment counts
    upcoming_count = upcoming_appointments.count()
    pending_count = pending_appointments.count()
    completed_count = completed_appointments.count()
    cancelled_count = cancelled_appointments.count()
    
    context = {
        'patient': patient,
        'upcoming_appointments': upcoming_appointments,
        'pending_appointments': pending_appointments,
        'completed_appointments': completed_page_obj,
        'cancelled_appointments': cancelled_appointments,
        'next_appointment': next_appointment,
        'upcoming_count': upcoming_count,
        'pending_count': pending_count,
        'completed_count': completed_count,
        'cancelled_count': cancelled_count,
    }
    return render(request, 'patients/appointments.html', context)


@login_required
def medical_history(request):
    """Patient medical history view with comprehensive data for all tabs"""
    if request.user.user_type != 'patient':
        messages.error(request, 'Access denied.')
        return redirect('healthcare_project:home')
    
    patient = get_object_or_404(Patient, user=request.user)
    
    # Get appointment history for Visit History tab
    appointments = Appointment.objects.filter(patient=patient).order_by('-appointment_date')
    
    # Create dummy medical records data (since models are disabled)
    medical_records = [
        {
            'date': '2025-09-18',
            'doctor': 'Dr. Sarah Johnson',
            'type': 'Annual Check-up',
            'notes': 'Routine physical examination. All vital signs normal.',
            'vitals': 'BP: 120/80, Weight: 68kg, Height: 170cm',
            'diagnosis': 'Healthy'
        },
        {
            'date': '2025-09-10',
            'doctor': 'Dr. Michael Brown',
            'type': 'Follow-up',
            'notes': 'Diabetes management review. Blood sugar levels stable.',
            'vitals': 'BP: 118/78, Weight: 67kg, HbA1c: 6.2%',
            'diagnosis': 'Type 2 Diabetes - Well Controlled'
        },
        {
            'date': '2025-08-15',
            'doctor': 'Dr. Emily Davis',
            'type': 'Specialist Consultation',
            'notes': 'Cardiology consultation for chest pain evaluation.',
            'vitals': 'BP: 125/82, HR: 72bpm, ECG: Normal',
            'diagnosis': 'Non-specific chest pain'
        },
        {
            'date': '2025-07-22',
            'doctor': 'Dr. Sarah Johnson',
            'type': 'Routine Visit',
            'notes': 'Regular check-up for diabetes monitoring.',
            'vitals': 'BP: 122/79, Weight: 67.5kg, HbA1c: 6.1%',
            'diagnosis': 'Type 2 Diabetes - Well Controlled'
        },
        {
            'date': '2025-06-30',
            'doctor': 'Dr. Lisa Wilson',
            'type': 'Dermatology Consultation',
            'notes': 'Skin examination for mole changes. No abnormalities found.',
            'vitals': 'BP: 119/76, Normal skin examination',
            'diagnosis': 'Healthy skin, continue routine monitoring'
        },
        {
            'date': '2025-06-15',
            'doctor': 'Dr. Michael Brown',
            'type': 'Follow-up',
            'notes': 'Blood pressure medication adjustment and review.',
            'vitals': 'BP: 128/84, Weight: 68kg, HR: 68bpm',
            'diagnosis': 'Hypertension - Adjusting medication'
        },
        {
            'date': '2025-05-20',
            'doctor': 'Dr. Robert Chen',
            'type': 'Eye Examination',
            'notes': 'Routine eye exam for diabetic patients. Retina healthy.',
            'vitals': 'Vision: 20/20, No diabetic retinopathy',
            'diagnosis': 'Healthy eyes'
        },
        {
            'date': '2025-04-18',
            'doctor': 'Dr. Sarah Johnson',
            'type': 'Physical Therapy',
            'notes': 'Knee rehabilitation assessment and exercise plan.',
            'vitals': 'Range of motion improved, mild swelling reduced',
            'diagnosis': 'Recovering well from knee injury'
        },
        {
            'date': '2025-03-25',
            'doctor': 'Dr. Amanda Foster',
            'type': 'Mental Health Check',
            'notes': 'Stress management and anxiety screening.',
            'vitals': 'BP: 121/77, Normal mental status exam',
            'diagnosis': 'Mild anxiety - lifestyle modifications recommended'
        },
        {
            'date': '2025-02-14',
            'doctor': 'Dr. Michael Brown',
            'type': 'Lab Review',
            'notes': 'Discussion of recent laboratory results and medication review.',
            'vitals': 'Weight: 68.2kg, All labs within normal limits',
            'diagnosis': 'Excellent progress'
        },
        {
            'date': '2025-01-30',
            'doctor': 'Dr. Sarah Johnson',
            'type': 'Preventive Care',
            'notes': 'Vaccination updates and health screening discussion.',
            'vitals': 'BP: 118/75, Weight: 68.5kg, Height: 170cm',
            'diagnosis': 'Up to date with preventive care'
        },
        {
            'date': '2025-01-10',
            'doctor': 'Dr. Jennifer Park',
            'type': 'Nutrition Consultation',
            'notes': 'Dietary counseling for diabetes management and weight control.',
            'vitals': 'Weight: 69kg, BMI: 23.9, Good nutritional status',
            'diagnosis': 'Nutritional goals on track'
        }
    ]
    
    # Create dummy prescriptions data
    prescriptions = [
        {
            'medication': 'Metformin',
            'dosage': '500mg',
            'frequency': 'Twice daily',
            'prescribed_date': '2025-09-10',
            'doctor': 'Dr. Michael Brown',
            'status': 'Active',
            'refills': 2,
            'instructions': 'Take with meals'
        },
        {
            'medication': 'Lisinopril',
            'dosage': '10mg',
            'frequency': 'Once daily',
            'prescribed_date': '2025-08-20',
            'doctor': 'Dr. Sarah Johnson',
            'status': 'Active',
            'refills': 3,
            'instructions': 'Take in the morning'
        },
        {
            'medication': 'Vitamin D3',
            'dosage': '1000 IU',
            'frequency': 'Once daily',
            'prescribed_date': '2025-09-18',
            'doctor': 'Dr. Sarah Johnson',
            'status': 'Active',
            'refills': 5,
            'instructions': 'Take with food'
        },
        {
            'medication': 'Ibuprofen',
            'dosage': '400mg',
            'frequency': 'As needed',
            'prescribed_date': '2025-07-10',
            'doctor': 'Dr. Emily Davis',
            'status': 'Completed',
            'refills': 0,
            'instructions': 'For pain relief, max 3 times daily'
        },
        {
            'medication': 'Atorvastatin',
            'dosage': '20mg',
            'frequency': 'Once daily',
            'prescribed_date': '2025-08-25',
            'doctor': 'Dr. Sarah Johnson',
            'status': 'Active',
            'refills': 5,
            'instructions': 'Take in the evening'
        },
        {
            'medication': 'Vitamin B12',
            'dosage': '1000mcg',
            'frequency': 'Once daily',
            'prescribed_date': '2025-04-08',
            'doctor': 'Dr. Jennifer Park',
            'status': 'Active',
            'refills': 3,
            'instructions': 'Take on empty stomach'
        },
        {
            'medication': 'Omeprazole',
            'dosage': '20mg',
            'frequency': 'Once daily',
            'prescribed_date': '2025-06-15',
            'doctor': 'Dr. Michael Brown',
            'status': 'Active',
            'refills': 2,
            'instructions': 'Take before breakfast'
        }
    ]
    
    # Create dummy lab results data
    lab_results = [
        {
            'test_name': 'Complete Blood Count (CBC)',
            'date': '2025-09-15',
            'lab': 'HealthLab Central',
            'results': 'All values within normal range',
            'status': 'Completed',
            'ordered_by': 'Dr. Sarah Johnson',
            'details': {
                'WBC': '7.2 K/µL (Normal: 4.5-11.0)',
                'RBC': '4.8 M/µL (Normal: 4.2-5.4)',
                'Hemoglobin': '14.2 g/dL (Normal: 12.0-16.0)',
                'Platelets': '285 K/µL (Normal: 150-450)'
            }
        },
        {
            'test_name': 'HbA1c (Diabetes Monitoring)',
            'date': '2025-09-05',
            'lab': 'HealthLab Central',
            'results': '6.2% (Good control)',
            'status': 'Completed',
            'ordered_by': 'Dr. Michael Brown',
            'details': {
                'HbA1c': '6.2% (Target: <7.0%)',
                'Glucose': '125 mg/dL (Fasting)'
            }
        },
        {
            'test_name': 'Lipid Panel',
            'date': '2025-08-25',
            'lab': 'Quest Diagnostics',
            'results': 'Cholesterol slightly elevated',
            'status': 'Completed',
            'ordered_by': 'Dr. Sarah Johnson',
            'details': {
                'Total Cholesterol': '215 mg/dL (Desirable: <200)',
                'LDL': '140 mg/dL (Optimal: <100)',
                'HDL': '45 mg/dL (Good: >40)',
                'Triglycerides': '150 mg/dL (Normal: <150)'
            }
        },
        {
            'test_name': 'Chest X-Ray',
            'date': '2025-08-15',
            'lab': 'Central Imaging',
            'results': 'Normal chest radiograph',
            'status': 'Completed',
            'ordered_by': 'Dr. Emily Davis',
            'details': {
                'Findings': 'No acute cardiopulmonary abnormality',
                'Impression': 'Normal chest X-ray'
            }
        },
        {
            'test_name': 'Comprehensive Metabolic Panel',
            'date': '2025-07-20',
            'lab': 'HealthLab Central',
            'results': 'All values normal except mild dehydration',
            'status': 'Completed',
            'ordered_by': 'Dr. Sarah Johnson',
            'details': {
                'Sodium': '142 mEq/L (Normal: 136-145)',
                'Potassium': '4.1 mEq/L (Normal: 3.5-5.0)',
                'Creatinine': '0.9 mg/dL (Normal: 0.6-1.2)',
                'BUN': '18 mg/dL (Normal: 7-20)'
            }
        },
        {
            'test_name': 'Thyroid Function Test',
            'date': '2025-06-10',
            'lab': 'Quest Diagnostics',
            'results': 'Normal thyroid function',
            'status': 'Completed',
            'ordered_by': 'Dr. Michael Brown',
            'details': {
                'TSH': '2.1 mIU/L (Normal: 0.4-4.0)',
                'T4': '7.8 µg/dL (Normal: 5.0-12.0)',
                'T3': '110 ng/dL (Normal: 80-200)'
            }
        },
        {
            'test_name': 'Urine Analysis',
            'date': '2025-05-15',
            'lab': 'HealthLab Central',
            'results': 'Normal urinalysis',
            'status': 'Completed',
            'ordered_by': 'Dr. Sarah Johnson',
            'details': {
                'Protein': 'Negative',
                'Glucose': 'Negative',
                'Ketones': 'Negative',
                'Specific Gravity': '1.015 (Normal: 1.005-1.030)'
            }
        },
        {
            'test_name': 'Vitamin B12 & Folate',
            'date': '2025-04-08',
            'lab': 'Quest Diagnostics',
            'results': 'Vitamin B12 slightly low, Folate normal',
            'status': 'Completed',
            'ordered_by': 'Dr. Jennifer Park',
            'details': {
                'Vitamin B12': '250 pg/mL (Normal: 300-900)',
                'Folate': '12 ng/mL (Normal: 3-20)',
                'Recommendation': 'B12 supplementation advised'
            }
        }
    ]
    
    # Create dummy documents data
    documents = [
        {
            'name': 'Insurance Card',
            'type': 'Insurance',
            'date_uploaded': '2025-01-15',
            'size': '245 KB',
            'format': 'PDF'
        },
        {
            'name': 'Vaccination Records',
            'type': 'Medical',
            'date_uploaded': '2025-03-20',
            'size': '180 KB',
            'format': 'PDF'
        },
        {
            'name': 'Allergy Test Results',
            'type': 'Test Results',
            'date_uploaded': '2025-08-12',
            'size': '320 KB',
            'format': 'PDF'
        },
        {
            'name': 'MRI Scan Report',
            'type': 'Imaging',
            'date_uploaded': '2025-09-05',
            'size': '2.1 MB',
            'format': 'PDF'
        },
        {
            'name': 'Blood Test Report - CBC',
            'type': 'Test Results',
            'date_uploaded': '2025-09-15',
            'size': '156 KB',
            'format': 'PDF'
        },
        {
            'name': 'Prescription History',
            'type': 'Medical',
            'date_uploaded': '2025-09-10',
            'size': '89 KB',
            'format': 'PDF'
        },
        {
            'name': 'Chest X-Ray Images',
            'type': 'Imaging',
            'date_uploaded': '2025-08-15',
            'size': '1.8 MB',
            'format': 'PDF'
        },
        {
            'name': 'Diabetes Management Plan',
            'type': 'Medical',
            'date_uploaded': '2025-07-22',
            'size': '95 KB',
            'format': 'PDF'
        },
        {
            'name': 'Emergency Contact Information',
            'type': 'Insurance',
            'date_uploaded': '2025-01-15',
            'size': '78 KB',
            'format': 'PDF'
        },
        {
            'name': 'Physical Therapy Assessment',
            'type': 'Medical',
            'date_uploaded': '2025-04-18',
            'size': '124 KB',
            'format': 'PDF'
        }
    ]
    
    # Calculate statistics for dashboard cards
    total_records = len(medical_records)
    active_prescriptions = len([p for p in prescriptions if p['status'] == 'Active'])
    recent_tests = len([r for r in lab_results if r['date'] >= '2025-08-01'])
    total_appointments = appointments.count()
    
    context = {
        'patient': patient,
        'appointments': appointments,
        'medical_records': medical_records,
        'prescriptions': prescriptions,
        'lab_results': lab_results,
        'documents': documents,
        'total_records': total_records,
        'active_prescriptions': active_prescriptions,
        'recent_tests': recent_tests,
        'total_appointments': total_appointments,
        # JSON-safe data for JavaScript
        'medical_records_json': json.dumps(medical_records),
        'prescriptions_json': json.dumps(prescriptions),
        'lab_results_json': json.dumps(lab_results),
        'documents_json': json.dumps(documents),
    }
    return render(request, 'patients/medical_records.html', context)


@login_required
def patient_documents(request):
    """Patient documents view"""
    if request.user.user_type != 'patient':
        messages.error(request, 'Access denied.')
        return redirect('healthcare_project:home')
    
    patient = get_object_or_404(Patient, user=request.user)
    
    # Get appointment history for context
    appointments = Appointment.objects.filter(patient=patient).order_by('-appointment_date')
    
    # Create comprehensive dummy medical records data
    medical_records = [
        {
            'date': '2025-09-18',
            'doctor': 'Dr. Sarah Johnson',
            'type': 'Annual Check-up',
            'notes': 'Routine physical examination. All vital signs normal.',
            'vitals': 'BP: 120/80, Weight: 68kg, Height: 170cm',
            'diagnosis': 'Healthy'
        },
        {
            'date': '2025-09-10',
            'doctor': 'Dr. Michael Brown',
            'type': 'Follow-up',
            'notes': 'Diabetes management review. Blood sugar levels stable.',
            'vitals': 'BP: 118/78, Weight: 67kg, HbA1c: 6.2%',
            'diagnosis': 'Type 2 Diabetes - Well Controlled'
        },
        {
            'date': '2025-08-15',
            'doctor': 'Dr. Emily Davis',
            'type': 'Specialist Consultation',
            'notes': 'Cardiology consultation for chest pain evaluation.',
            'vitals': 'BP: 125/82, HR: 72bpm, ECG: Normal',
            'diagnosis': 'Non-specific chest pain'
        },
        {
            'date': '2025-07-22',
            'doctor': 'Dr. Sarah Johnson',
            'type': 'Routine Visit',
            'notes': 'Regular check-up for diabetes monitoring.',
            'vitals': 'BP: 122/79, Weight: 67.5kg, HbA1c: 6.1%',
            'diagnosis': 'Type 2 Diabetes - Well Controlled'
        },
        {
            'date': '2025-06-30',
            'doctor': 'Dr. Lisa Wilson',
            'type': 'Dermatology Consultation',
            'notes': 'Skin examination for mole changes. No abnormalities found.',
            'vitals': 'BP: 119/76, Normal skin examination',
            'diagnosis': 'Healthy skin, continue routine monitoring'
        },
        {
            'date': '2025-06-15',
            'doctor': 'Dr. Michael Brown',
            'type': 'Follow-up',
            'notes': 'Blood pressure medication adjustment and review.',
            'vitals': 'BP: 128/84, Weight: 68kg, HR: 68bpm',
            'diagnosis': 'Hypertension - Adjusting medication'
        },
        {
            'date': '2025-05-20',
            'doctor': 'Dr. Robert Chen',
            'type': 'Eye Examination',
            'notes': 'Routine eye exam for diabetic patients. Retina healthy.',
            'vitals': 'Vision: 20/20, No diabetic retinopathy',
            'diagnosis': 'Healthy eyes'
        },
        {
            'date': '2025-04-18',
            'doctor': 'Dr. Sarah Johnson',
            'type': 'Physical Therapy',
            'notes': 'Knee rehabilitation assessment and exercise plan.',
            'vitals': 'Range of motion improved, mild swelling reduced',
            'diagnosis': 'Recovering well from knee injury'
        },
        {
            'date': '2025-03-25',
            'doctor': 'Dr. Amanda Foster',
            'type': 'Mental Health Check',
            'notes': 'Stress management and anxiety screening.',
            'vitals': 'BP: 121/77, Normal mental status exam',
            'diagnosis': 'Mild anxiety - lifestyle modifications recommended'
        },
        {
            'date': '2025-02-14',
            'doctor': 'Dr. Michael Brown',
            'type': 'Lab Review',
            'notes': 'Discussion of recent laboratory results and medication review.',
            'vitals': 'Weight: 68.2kg, All labs within normal limits',
            'diagnosis': 'Excellent progress'
        },
        {
            'date': '2025-01-30',
            'doctor': 'Dr. Sarah Johnson',
            'type': 'Preventive Care',
            'notes': 'Vaccination updates and health screening discussion.',
            'vitals': 'BP: 118/75, Weight: 68.5kg, Height: 170cm',
            'diagnosis': 'Up to date with preventive care'
        },
        {
            'date': '2025-01-10',
            'doctor': 'Dr. Jennifer Park',
            'type': 'Nutrition Consultation',
            'notes': 'Dietary counseling for diabetes management and weight control.',
            'vitals': 'Weight: 69kg, BMI: 23.9, Good nutritional status',
            'diagnosis': 'Nutritional goals on track'
        }
    ]

    # Create comprehensive prescriptions data
    prescriptions = [
        {
            'medication': 'Metformin',
            'dosage': '500mg',
            'frequency': 'Twice daily',
            'prescribed_date': '2025-09-10',
            'doctor': 'Dr. Michael Brown',
            'status': 'Active',
            'refills': 2,
            'instructions': 'Take with meals'
        },
        {
            'medication': 'Lisinopril',
            'dosage': '10mg',
            'frequency': 'Once daily',
            'prescribed_date': '2025-08-20',
            'doctor': 'Dr. Sarah Johnson',
            'status': 'Active',
            'refills': 3,
            'instructions': 'Take in the morning'
        },
        {
            'medication': 'Vitamin D3',
            'dosage': '1000 IU',
            'frequency': 'Once daily',
            'prescribed_date': '2025-09-18',
            'doctor': 'Dr. Sarah Johnson',
            'status': 'Active',
            'refills': 5,
            'instructions': 'Take with food'
        },
        {
            'medication': 'Ibuprofen',
            'dosage': '400mg',
            'frequency': 'As needed',
            'prescribed_date': '2025-07-10',
            'doctor': 'Dr. Emily Davis',
            'status': 'Completed',
            'refills': 0,
            'instructions': 'For pain relief, max 3 times daily'
        },
        {
            'medication': 'Atorvastatin',
            'dosage': '20mg',
            'frequency': 'Once daily',
            'prescribed_date': '2025-08-25',
            'doctor': 'Dr. Sarah Johnson',
            'status': 'Active',
            'refills': 5,
            'instructions': 'Take in the evening'
        },
        {
            'medication': 'Vitamin B12',
            'dosage': '1000mcg',
            'frequency': 'Once daily',
            'prescribed_date': '2025-04-08',
            'doctor': 'Dr. Jennifer Park',
            'status': 'Active',
            'refills': 3,
            'instructions': 'Take on empty stomach'
        },
        {
            'medication': 'Omeprazole',
            'dosage': '20mg',
            'frequency': 'Once daily',
            'prescribed_date': '2025-06-15',
            'doctor': 'Dr. Michael Brown',
            'status': 'Active',
            'refills': 2,
            'instructions': 'Take before breakfast'
        }
    ]

    # Create comprehensive lab results data
    lab_results = [
        {
            'test_name': 'Complete Blood Count (CBC)',
            'date': '2025-09-15',
            'lab': 'HealthLab Central',
            'results': 'All values within normal range',
            'status': 'Completed',
            'ordered_by': 'Dr. Sarah Johnson',
            'details': {
                'WBC': '7.2 K/µL (Normal: 4.5-11.0)',
                'RBC': '4.8 M/µL (Normal: 4.2-5.4)',
                'Hemoglobin': '14.2 g/dL (Normal: 12.0-16.0)',
                'Platelets': '285 K/µL (Normal: 150-450)'
            }
        },
        {
            'test_name': 'HbA1c (Diabetes Monitoring)',
            'date': '2025-09-05',
            'lab': 'HealthLab Central',
            'results': '6.2% (Good control)',
            'status': 'Completed',
            'ordered_by': 'Dr. Michael Brown',
            'details': {
                'HbA1c': '6.2% (Target: <7.0%)',
                'Glucose': '125 mg/dL (Fasting)'
            }
        },
        {
            'test_name': 'Lipid Panel',
            'date': '2025-08-25',
            'lab': 'Quest Diagnostics',
            'results': 'Cholesterol slightly elevated',
            'status': 'Completed',
            'ordered_by': 'Dr. Sarah Johnson',
            'details': {
                'Total Cholesterol': '215 mg/dL (Desirable: <200)',
                'LDL': '140 mg/dL (Optimal: <100)',
                'HDL': '45 mg/dL (Good: >40)',
                'Triglycerides': '150 mg/dL (Normal: <150)'
            }
        },
        {
            'test_name': 'Chest X-Ray',
            'date': '2025-08-15',
            'lab': 'Central Imaging',
            'results': 'Normal chest radiograph',
            'status': 'Completed',
            'ordered_by': 'Dr. Emily Davis',
            'details': {
                'Findings': 'No acute cardiopulmonary abnormality',
                'Impression': 'Normal chest X-ray'
            }
        },
        {
            'test_name': 'Comprehensive Metabolic Panel',
            'date': '2025-07-20',
            'lab': 'HealthLab Central',
            'results': 'All values normal except mild dehydration',
            'status': 'Completed',
            'ordered_by': 'Dr. Sarah Johnson',
            'details': {
                'Sodium': '142 mEq/L (Normal: 136-145)',
                'Potassium': '4.1 mEq/L (Normal: 3.5-5.0)',
                'Creatinine': '0.9 mg/dL (Normal: 0.6-1.2)',
                'BUN': '18 mg/dL (Normal: 7-20)'
            }
        },
        {
            'test_name': 'Thyroid Function Test',
            'date': '2025-06-10',
            'lab': 'Quest Diagnostics',
            'results': 'Normal thyroid function',
            'status': 'Completed',
            'ordered_by': 'Dr. Michael Brown',
            'details': {
                'TSH': '2.1 mIU/L (Normal: 0.4-4.0)',
                'T4': '7.8 µg/dL (Normal: 5.0-12.0)',
                'T3': '110 ng/dL (Normal: 80-200)'
            }
        },
        {
            'test_name': 'Urine Analysis',
            'date': '2025-05-15',
            'lab': 'HealthLab Central',
            'results': 'Normal urinalysis',
            'status': 'Completed',
            'ordered_by': 'Dr. Sarah Johnson',
            'details': {
                'Protein': 'Negative',
                'Glucose': 'Negative',
                'Ketones': 'Negative',
                'Specific Gravity': '1.015 (Normal: 1.005-1.030)'
            }
        },
        {
            'test_name': 'Vitamin B12 & Folate',
            'date': '2025-04-08',
            'lab': 'Quest Diagnostics',
            'results': 'Vitamin B12 slightly low, Folate normal',
            'status': 'Completed',
            'ordered_by': 'Dr. Jennifer Park',
            'details': {
                'Vitamin B12': '250 pg/mL (Normal: 300-900)',
                'Folate': '12 ng/mL (Normal: 3-20)',
                'Recommendation': 'B12 supplementation advised'
            }
        }
    ]

    # Create comprehensive documents data
    documents = [
        {
            'name': 'Insurance Card',
            'type': 'Insurance',
            'date_uploaded': '2025-01-15',
            'size': '245 KB',
            'format': 'PDF'
        },
        {
            'name': 'Vaccination Records',
            'type': 'Medical',
            'date_uploaded': '2025-03-20',
            'size': '180 KB',
            'format': 'PDF'
        },
        {
            'name': 'Allergy Test Results',
            'type': 'Test Results',
            'date_uploaded': '2025-08-12',
            'size': '320 KB',
            'format': 'PDF'
        },
        {
            'name': 'MRI Scan Report',
            'type': 'Imaging',
            'date_uploaded': '2025-09-05',
            'size': '2.1 MB',
            'format': 'PDF'
        },
        {
            'name': 'Blood Test Report - CBC',
            'type': 'Test Results',
            'date_uploaded': '2025-09-15',
            'size': '156 KB',
            'format': 'PDF'
        },
        {
            'name': 'Prescription History',
            'type': 'Medical',
            'date_uploaded': '2025-09-10',
            'size': '89 KB',
            'format': 'PDF'
        },
        {
            'name': 'Chest X-Ray Images',
            'type': 'Imaging',
            'date_uploaded': '2025-08-15',
            'size': '1.8 MB',
            'format': 'PDF'
        },
        {
            'name': 'Diabetes Management Plan',
            'type': 'Medical',
            'date_uploaded': '2025-07-22',
            'size': '95 KB',
            'format': 'PDF'
        },
        {
            'name': 'Emergency Contact Information',
            'type': 'Insurance',
            'date_uploaded': '2025-01-15',
            'size': '78 KB',
            'format': 'PDF'
        },
        {
            'name': 'Physical Therapy Assessment',
            'type': 'Medical',
            'date_uploaded': '2025-04-18',
            'size': '124 KB',
            'format': 'PDF'
        }
    ]
    
    # Calculate statistics for dashboard cards
    total_records = len(medical_records)
    active_prescriptions = len([p for p in prescriptions if p['status'] == 'Active'])
    recent_tests = len([r for r in lab_results if r['date'] >= '2025-08-01'])
    total_appointments = appointments.count()
    
    # Calculate document type counts
    imaging_docs = len([d for d in documents if d['type'] == 'Imaging'])
    insurance_docs = len([d for d in documents if d['type'] == 'Insurance'])
    medical_docs = len([d for d in documents if d['type'] == 'Medical'])
    test_result_docs = len([d for d in documents if d['type'] == 'Test Results'])

    context = {
        'patient': patient,
        'appointments': appointments,
        'medical_records': medical_records,
        'prescriptions': prescriptions,
        'lab_results': lab_results,
        'documents': documents,
        'total_records': total_records,
        'active_prescriptions': active_prescriptions,
        'recent_tests': recent_tests,
        'total_appointments': total_appointments,
        'imaging_docs': imaging_docs,
        'insurance_docs': insurance_docs,
        'medical_docs': medical_docs,
        'test_result_docs': test_result_docs,
        # JSON-safe data for JavaScript
        'prescriptions_json': json.dumps(prescriptions),
        'lab_results_json': json.dumps(lab_results),
        'documents_json': json.dumps(documents),
    }
    return render(request, 'patients/documents.html', context)


@login_required
def book_appointment(request):
    """Book new appointment view"""
    if request.user.user_type != 'patient':
        messages.error(request, 'Access denied.')
        return redirect('healthcare_project:home')
    
    patient = get_object_or_404(Patient, user=request.user)
    
    if request.method == 'POST':
        # Handle appointment booking
        doctor_id = request.POST.get('doctor')
        appointment_date = request.POST.get('appointment_date')
        appointment_time = request.POST.get('appointment_time')
        reason = request.POST.get('reason')
        appointment_type = request.POST.get('appointment_type')
        
        try:
            doctor = Doctor.objects.get(id=doctor_id)
            
            appointment = Appointment.objects.create(
                patient=patient,
                doctor=doctor,
                appointment_date=appointment_date,
                appointment_time=appointment_time,
                reason=reason,
                appointment_type=appointment_type,
                status='pending'
            )
            
            messages.success(request, 'Appointment request submitted successfully. You will receive confirmation shortly.')
            return redirect('patients:appointments')
            
        except Doctor.DoesNotExist:
            messages.error(request, 'Selected doctor not found.')
        except Exception as e:
            messages.error(request, f'Error booking appointment: {str(e)}')
    
    # Get available doctors
    doctors = Doctor.objects.all().select_related('user')
    
    context = {
        'patient': patient,
        'doctors': doctors,
    }
    return render(request, 'patients/book_appointment.html', context)


@login_required
def export_medical_records(request):
    """Export patient medical records"""
    if request.user.user_type != 'patient':
        messages.error(request, 'Access denied.')
        return redirect('healthcare_project:home')
    
    if request.method != 'POST':
        return redirect('patients:medical_history')
    
    patient = get_object_or_404(Patient, user=request.user)
    
    import json
    from django.http import HttpResponse
    from datetime import datetime, timedelta
    import io
    
    try:
        # Get form data
        record_types = json.loads(request.POST.get('record_types', '[]'))
        date_range = request.POST.get('date_range', 'all')
        export_format = request.POST.get('format', 'pdf')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        
        # Determine date filtering
        today = datetime.now().date()
        if date_range == 'last_year':
            filter_start = today - timedelta(days=365)
            filter_end = today
        elif date_range == 'last_6months':
            filter_start = today - timedelta(days=180)
            filter_end = today
        elif date_range == 'last_3months':
            filter_start = today - timedelta(days=90)
            filter_end = today
        elif date_range == 'custom' and start_date and end_date:
            filter_start = datetime.strptime(start_date, '%Y-%m-%d').date()
            filter_end = datetime.strptime(end_date, '%Y-%m-%d').date()
        else:
            filter_start = None
            filter_end = None
        
        # Collect data based on selected types
        data = {}
        
        if 'appointments' in record_types:
            appointments = Appointment.objects.filter(patient=patient)
            if filter_start and filter_end:
                appointments = appointments.filter(
                    appointment_date__gte=filter_start,
                    appointment_date__lte=filter_end
                )
            data['appointments'] = appointments.order_by('-appointment_date')
        
        # For now, create dummy medical records since the models are disabled
        if 'medical_records' in record_types:
            data['medical_records'] = [
                {
                    'date': '2024-12-21',
                    'doctor': 'Dr. Sarah Johnson',
                    'type': 'Annual Check-up',
                    'notes': 'Routine physical examination. All vital signs normal.',
                    'vitals': 'BP: 120/80, Weight: 68kg, Height: 170cm'
                },
                {
                    'date': '2024-11-15',
                    'doctor': 'Dr. Michael Brown',
                    'type': 'Follow-up',
                    'notes': 'Monitoring blood pressure medication effectiveness.',
                    'vitals': 'BP: 125/85, Weight: 69kg'
                }
            ]
        
        if 'prescriptions' in record_types:
            data['prescriptions'] = [
                {
                    'medication': 'Lisinopril',
                    'dosage': '10mg',
                    'frequency': 'Once daily',
                    'prescribed_date': '2024-11-15',
                    'doctor': 'Dr. Sarah Johnson',
                    'status': 'Active'
                },
                {
                    'medication': 'Vitamin D3',
                    'dosage': '2000 IU',
                    'frequency': 'Once daily',
                    'prescribed_date': '2024-10-01',
                    'doctor': 'Dr. Sarah Johnson',
                    'status': 'Active'
                }
            ]
        
        if 'lab_results' in record_types:
            data['lab_results'] = [
                {
                    'test_name': 'Complete Blood Count',
                    'date': '2024-12-18',
                    'lab': 'City Medical Lab',
                    'results': 'All values within normal range',
                    'status': 'Normal'
                },
                {
                    'test_name': 'Lipid Panel',
                    'date': '2024-12-18',
                    'lab': 'City Medical Lab',
                    'results': 'Cholesterol slightly elevated',
                    'status': 'Monitor'
                }
            ]
        
        # Generate export based on format
        if export_format == 'pdf':
            return generate_pdf_report(patient, data, filter_start, filter_end)
        elif export_format == 'excel':
            return generate_excel_report(patient, data, filter_start, filter_end)
        elif export_format == 'csv':
            return generate_csv_report(patient, data, filter_start, filter_end)
        
    except Exception as e:
        messages.error(request, f'Export failed: {str(e)}')
        return redirect('patients:medical_history')


def generate_pdf_report(patient, data, start_date, end_date):
    """Generate PDF medical records report"""
    from django.http import HttpResponse
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    import io
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title = Paragraph(f"Medical Records Report - {patient.user.get_full_name()}", styles['Title'])
    story.append(title)
    story.append(Spacer(1, 12))
    
    # Date range
    if start_date and end_date:
        date_info = Paragraph(f"Date Range: {start_date} to {end_date}", styles['Normal'])
        story.append(date_info)
        story.append(Spacer(1, 12))
    
    # Patient Information
    patient_info = Paragraph(f"""
        <b>Patient Information:</b><br/>
        Name: {patient.user.get_full_name()}<br/>
        Date of Birth: {getattr(patient, 'date_of_birth', 'Not specified')}<br/>
        Phone: {getattr(patient, 'phone', 'Not specified')}<br/>
        Email: {patient.user.email}<br/>
        Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}
    """, styles['Normal'])
    story.append(patient_info)
    story.append(Spacer(1, 12))
    
    # Add each data section
    for section, items in data.items():
        if section == 'appointments' and items:
            story.append(Paragraph('<b>Appointments</b>', styles['Heading2']))
            table_data = [['Date', 'Time', 'Doctor', 'Type', 'Status']]
            for apt in items:
                table_data.append([
                    str(apt.appointment_date),
                    str(apt.appointment_time),
                    f"Dr. {apt.doctor.user.get_full_name()}",
                    apt.get_appointment_type_display(),
                    apt.get_status_display()
                ])
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)
            story.append(Spacer(1, 12))
        
        elif section == 'medical_records' and items:
            story.append(Paragraph('<b>Medical Records</b>', styles['Heading2']))
            for record in items:
                record_p = Paragraph(f"""
                    <b>Date:</b> {record['date']}<br/>
                    <b>Doctor:</b> {record['doctor']}<br/>
                    <b>Type:</b> {record['type']}<br/>
                    <b>Notes:</b> {record['notes']}<br/>
                    <b>Vitals:</b> {record['vitals']}<br/><br/>
                """, styles['Normal'])
                story.append(record_p)
            story.append(Spacer(1, 12))
        
        elif section == 'prescriptions' and items:
            story.append(Paragraph('<b>Prescriptions</b>', styles['Heading2']))
            table_data = [['Medication', 'Dosage', 'Frequency', 'Date', 'Doctor', 'Status']]
            for rx in items:
                table_data.append([
                    rx['medication'],
                    rx['dosage'],
                    rx['frequency'],
                    rx['prescribed_date'],
                    rx['doctor'],
                    rx['status']
                ])
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)
            story.append(Spacer(1, 12))
        
        elif section == 'lab_results' and items:
            story.append(Paragraph('<b>Lab Results</b>', styles['Heading2']))
            table_data = [['Test', 'Date', 'Lab', 'Results', 'Status']]
            for lab in items:
                table_data.append([
                    lab['test_name'],
                    lab['date'],
                    lab['lab'],
                    lab['results'],
                    lab['status']
                ])
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="medical_records_{patient.user.username}_{datetime.now().strftime("%Y%m%d")}.pdf"'
    return response


def generate_excel_report(patient, data, start_date, end_date):
    """Generate Excel medical records report"""
    import openpyxl
    from django.http import HttpResponse
    import io
    
    wb = openpyxl.Workbook()
    
    # Remove default sheet and add our sheets
    wb.remove(wb.active)
    
    # Patient Info Sheet
    info_sheet = wb.create_sheet("Patient Information")
    info_sheet['A1'] = "Patient Medical Records Export"
    info_sheet['A1'].font = openpyxl.styles.Font(bold=True, size=16)
    info_sheet['A3'] = "Name:"
    info_sheet['B3'] = patient.user.get_full_name()
    info_sheet['A4'] = "Email:"
    info_sheet['B4'] = patient.user.email
    info_sheet['A5'] = "Export Date:"
    info_sheet['B5'] = datetime.now().strftime('%Y-%m-%d %H:%M')
    
    if start_date and end_date:
        info_sheet['A6'] = "Date Range:"
        info_sheet['B6'] = f"{start_date} to {end_date}"
    
    # Add data sheets
    for section, items in data.items():
        if section == 'appointments' and items:
            sheet = wb.create_sheet("Appointments")
            headers = ['Date', 'Time', 'Doctor', 'Type', 'Status', 'Notes']
            for col, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col).value = header
                sheet.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
            
            for row, apt in enumerate(items, 2):
                sheet.cell(row=row, column=1).value = str(apt.appointment_date)
                sheet.cell(row=row, column=2).value = str(apt.appointment_time)
                sheet.cell(row=row, column=3).value = f"Dr. {apt.doctor.user.get_full_name()}"
                sheet.cell(row=row, column=4).value = apt.get_appointment_type_display()
                sheet.cell(row=row, column=5).value = apt.get_status_display()
                sheet.cell(row=row, column=6).value = getattr(apt, 'notes', '')
        
        elif section == 'prescriptions' and items:
            sheet = wb.create_sheet("Prescriptions")
            headers = ['Medication', 'Dosage', 'Frequency', 'Date Prescribed', 'Doctor', 'Status']
            for col, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col).value = header
                sheet.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
            
            for row, rx in enumerate(items, 2):
                sheet.cell(row=row, column=1).value = rx['medication']
                sheet.cell(row=row, column=2).value = rx['dosage']
                sheet.cell(row=row, column=3).value = rx['frequency']
                sheet.cell(row=row, column=4).value = rx['prescribed_date']
                sheet.cell(row=row, column=5).value = rx['doctor']
                sheet.cell(row=row, column=6).value = rx['status']
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="medical_records_{patient.user.username}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    return response


def generate_csv_report(patient, data, start_date, end_date):
    """Generate CSV medical records report"""
    import csv
    from django.http import HttpResponse
    import io
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="medical_records_{patient.user.username}_{datetime.now().strftime("%Y%m%d")}.csv"'
    
    writer = csv.writer(response)
    
    # Header information
    writer.writerow(['Medical Records Export'])
    writer.writerow(['Patient:', patient.user.get_full_name()])
    writer.writerow(['Email:', patient.user.email])
    writer.writerow(['Export Date:', datetime.now().strftime('%Y-%m-%d %H:%M')])
    if start_date and end_date:
        writer.writerow(['Date Range:', f"{start_date} to {end_date}"])
    writer.writerow([])  # Empty row
    
    # Add data sections
    for section, items in data.items():
        if section == 'appointments' and items:
            writer.writerow(['APPOINTMENTS'])
            writer.writerow(['Date', 'Time', 'Doctor', 'Type', 'Status'])
            for apt in items:
                writer.writerow([
                    str(apt.appointment_date),
                    str(apt.appointment_time),
                    f"Dr. {apt.doctor.user.get_full_name()}",
                    apt.get_appointment_type_display(),
                    apt.get_status_display()
                ])
            writer.writerow([])  # Empty row
        
        elif section == 'prescriptions' and items:
            writer.writerow(['PRESCRIPTIONS'])
            writer.writerow(['Medication', 'Dosage', 'Frequency', 'Date Prescribed', 'Doctor', 'Status'])
            for rx in items:
                writer.writerow([
                    rx['medication'],
                    rx['dosage'],
                    rx['frequency'],
                    rx['prescribed_date'],
                    rx['doctor'],
                    rx['status']
                ])
            writer.writerow([])  # Empty row
    
    return response


@login_required
def request_medical_records(request):
    """Handle medical records requests from patients"""
    if request.method == 'POST':
        try:
            # Get form data
            request_type = request.POST.get('request_type')
            provider_name = request.POST.get('provider_name')
            provider_contact = request.POST.get('provider_contact', '')
            date_from = request.POST.get('date_from')
            date_to = request.POST.get('date_to')
            details = request.POST.get('details', '')
            urgency = request.POST.get('urgency', 'routine')
            
            # Validate required fields
            if not request_type or not provider_name:
                return JsonResponse({'error': 'Missing required fields'}, status=400)
            
            # For now, we'll just log the request and return success
            # In a real system, this would create a database record and possibly send emails
            import logging
            logger = logging.getLogger(__name__)
            logger.info(f"Medical records request from user {request.user.id}: {request_type} from {provider_name}")
            
            # You could add this to a MedicalRecordRequest model:
            # MedicalRecordRequest.objects.create(
            #     patient=request.user.patient,
            #     request_type=request_type,
            #     provider_name=provider_name,
            #     provider_contact=provider_contact,
            #     date_from=date_from if date_from else None,
            #     date_to=date_to if date_to else None,
            #     details=details,
            #     urgency=urgency,
            #     status='pending',
            #     created_at=timezone.now()
            # )
            
            return JsonResponse({
                'success': True,
                'message': 'Request submitted successfully'
            })
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Invalid request method'}, status=405)


@login_required
def view_document(request, document_name):
    """View document details"""
    if request.user.user_type != 'patient':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    patient = get_object_or_404(Patient, user=request.user)
    
    # Get document from our dummy data (in real app, this would be from database)
    documents = [
        {
            'name': 'Insurance Card',
            'type': 'Insurance',
            'date_uploaded': '2024-01-15',
            'size': '245 KB',
            'format': 'PDF',
            'description': 'Current health insurance card with coverage details',
            'provider': 'Blue Cross Blue Shield',
            'policy_number': 'BC123456789'
        },
        {
            'name': 'Vaccination Records',
            'type': 'Medical',
            'date_uploaded': '2025-03-20',
            'size': '180 KB',
            'format': 'PDF',
            'description': 'Complete vaccination history including COVID-19, flu, and routine immunizations',
            'provider': 'City Health Department',
            'notes': 'Up to date with all recommended vaccinations'
        },
        {
            'name': 'Allergy Test Results',
            'type': 'Test Results',
            'date_uploaded': '2025-08-12',
            'size': '320 KB',
            'format': 'PDF',
            'description': 'Comprehensive allergy panel testing results',
            'provider': 'Allergy & Asthma Clinic',
            'findings': 'Mild allergies to pollen and dust mites'
        },
        {
            'name': 'MRI Scan Report',
            'type': 'Imaging',
            'date_uploaded': '2025-09-05',
            'size': '2.1 MB',
            'format': 'PDF',
            'description': 'Brain MRI scan with radiologist interpretation',
            'provider': 'Central Imaging Center',
            'findings': 'No abnormalities detected'
        }
    ]
    
    # Find the requested document
    document = next((doc for doc in documents if doc['name'] == document_name), None)
    if not document:
        return JsonResponse({'error': 'Document not found'}, status=404)
    
    return JsonResponse({'success': True, 'document': document})


@login_required
def download_document(request, document_name):
    """Download document file"""
    if request.user.user_type != 'patient':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    patient = get_object_or_404(Patient, user=request.user)
    
    # In a real application, you would:
    # 1. Verify the document belongs to this patient
    # 2. Get the actual file from storage
    # 3. Return the file as a download
    
    # For now, we'll simulate a download by creating a dummy PDF
    from django.http import HttpResponse
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter
    import io
    
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    
    # Add document content
    p.drawString(100, 750, f"Healthcare Management System")
    p.drawString(100, 720, f"Patient Document: {document_name}")
    p.drawString(100, 690, f"Patient: {patient.user.get_full_name()}")
    p.drawString(100, 660, f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    p.drawString(100, 600, f"This is a sample document for: {document_name}")
    p.drawString(100, 570, f"In a real system, this would be the actual document file.")
    
    p.showPage()
    p.save()
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{document_name.replace(" ", "_")}.pdf"'
    return response


@login_required
def upload_document(request):
    """Handle document upload"""
    if request.user.user_type != 'patient':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    if request.method == 'POST':
        try:
            # Get form data
            document_name = request.POST.get('document_name')
            document_type = request.POST.get('document_type')
            document_file = request.FILES.get('document_file')
            description = request.POST.get('description', '')
            
            if not document_name or not document_type or not document_file:
                return JsonResponse({'error': 'Missing required fields'}, status=400)
            
            # In a real application, you would:
            # 1. Save the file to storage (filesystem, S3, etc.)
            # 2. Create a database record
            # 3. Validate file type and size
            
            # For now, we'll just simulate successful upload
            return JsonResponse({
                'success': True,
                'message': f'Document "{document_name}" uploaded successfully',
                'document': {
                    'name': document_name,
                    'type': document_type,
                    'size': f'{document_file.size / 1024:.0f} KB',
                    'format': document_file.name.split('.')[-1].upper(),
                    'date_uploaded': datetime.now().strftime('%Y-%m-%d')
                }
            })
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Invalid request method'}, status=405)


@login_required
def patient_appointment_detail(request, appointment_id):
    """Patient appointment detail view - allows patients to view details of their own appointments"""
    if request.user.user_type != 'patient':
        messages.error(request, 'Access denied.')
        return redirect('healthcare_project:home')
    
    patient = get_object_or_404(Patient, user=request.user)
    appointment = get_object_or_404(Appointment, id=appointment_id, patient=patient)
    
    context = {
        'appointment': appointment,
        'patient': patient,
    }
    return render(request, 'patients/appointment_detail.html', context)


# Dashboard view
